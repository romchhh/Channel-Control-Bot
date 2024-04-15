import sqlite3
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.workbook import Workbook


def get_DB():
    infos = []
    conn = sqlite3.connect('bot.db')
    c = conn.cursor()
    c.execute("SELECT user_id, name, first_name, email FROM clients")

    info = c.fetchone()
    while info is not None:
        infos.append(info)
        info = c.fetchone()

    wb = Workbook()
    ws = wb.active

    header = ['ID', 'Username', 'First Name', 'Email']
    ws.append(header)

    for row_data in infos:
        ws.append(row_data)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(header)):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 10)
        ws.column_dimensions[column].width = adjusted_width

    file_name = 'xlsx/clients_info.xlsx'
    wb.save(file_name)
    doc = open(file_name, 'rb')
    return doc


